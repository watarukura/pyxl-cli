import csv
from typing import Any, List

from openpyxl import load_workbook, worksheet
from openpyxl.utils.cell import (
    column_index_from_string, coordinate_from_string
)


def write_xlsx(
    template_xlsx: str,
    sheet_xy_csv: List[Any],
    output_xlsx: str,
    delimiter=",",
):
    """Xlsxファイルへの書き込みを行う

    Args:
        template_xlsx (str): 入力Xlsxファイル名
        sheet_xy_csv (List[Any]): シートNo + セル位置 ＋ 入力ファイル
        output_xlsx (str): 出力Xlsxファイル名
    """
    wb = load_workbook(filename=template_xlsx)

    for sheet_no, address, input_file in sheet_xy_csv:
        with open(input_file) as f:
            reader = csv.reader(f, delimiter=delimiter)
            input_csv = [row for row in reader]
            ws = wb.worksheets[sheet_no - 1]
            start_col_letter, start_row = coordinate_from_string(address)
            start_col = column_index_from_string(start_col_letter)
            write_list_2d(ws, input_csv, start_row, start_col)

    wb.save(output_xlsx)


def write_list_2d(
    sheet: worksheet, l_2d: List[List[str]], start_row: int, start_col: int,
) -> None:
    """2次元配列を指定したセル位置から書き込む
    https://github.com/nkmk/python-snippets/blob/507a7fcb5b0a212d88015ffad77914c50509d4bc/notebook/openpyxl_example.py#L99-L116

    Args:
        sheet (sheet): Xlsxのsheetオブジェクト
        l_2d (list): 2次元配列
        start_row (int): 行(y)開始位置
        start_col (int): 列(x)開始位置
    """
    for y, row in enumerate(l_2d):
        for x, cell in enumerate(row):
            sheet.cell(
                row=start_row + y, column=start_col + x, value=l_2d[y][x]
            )
