from typing import Tuple

from openpyxl import coordinate_from_string, load_workbook, worksheet


def write_xlsx(
    template_xlsx: str,
    sheet_no: int,
    address: str,
    input_csv: Tuple[Tuple[str, int]],
    output_xlsx: str,
):
    """Xlsxファイルへの書き込みを行う

    Args:
        template_xlsx (str): 入力Xlsxファイル名
        sheet_no (int): sheet number(1始まり)
        address (str): 開始セル位置(ex: A1)
        input_csv (Tuple[Tuple[str, int]]): 書き込み内容の2次元配列
        output_xlsx (str): 出力Xlsxファイル名
    """
    wb = load_workbook(filename=template_xlsx)
    sheet_name = wb.get_sheet_names()[sheet_no - 1]
    ws = wb.get_sheet_by_name(sheet_name)

    start_row, start_col = coordinate_from_string(address)
    write_list_2d(ws, input_csv, start_row, start_col)

    wb.save(output_xlsx)


def write_list_2d(
    sheet: worksheet,
    l_2d: Tuple[Tuple[str, int]],
    start_row: int,
    start_col: int,
) -> None:
    """2次元配列を指定したセル位置から書き込む
    https://github.com/nkmk/python-snippets/blob/507a7fcb5b0a212d88015ffad77914c50509d4bc/notebook/openpyxl_example.py#L99-L116

    Args:
        sheet (sheet): Xlsxのsheetオブジェクト
        l_2d (list): 2次元配列
        start_row (int): 行(y)開始位置
        start_col (int): 列(x)開始位置
    """
    for y, row in enumerate(l_2d):
        for x, cell in enumerate(row):
            sheet.cell(
                row=start_row + y, column=start_col + x, value=l_2d[y][x]
            )
